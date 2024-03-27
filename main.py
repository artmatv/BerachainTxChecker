import requests
import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook("wallets.xlsx")

# Get the active worksheet
worksheet = workbook.active

# Read the addresses from a specific column (e.g., column A)
addresses = [str(cell.value) for cell in worksheet["A"][1:] if cell.value]

nodeUrl="https://rpc.ankr.com/berachain_testnet"

MAX_BATCH_SIZE = 100

def GetAccountTx(addresses):
    results = []
    batches = [addresses[i:i + MAX_BATCH_SIZE] for i in range(0, len(addresses), MAX_BATCH_SIZE)]

    for batch in batches:
        request_data = generate_request_data(batch)
        response = requests.post(nodeUrl, json=request_data, headers={"Content-Type": "application/json"})
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list):
                results.extend([int(entry["result"], 16) for entry in data])
            else:
                raise Exception("Unexpected response format")
        else:
            raise Exception(f"Error: {response.status_code}")

    return results

def generate_request_data(batch_addresses):
    request_array = []
    for index, address in enumerate(batch_addresses):
        request_id = index + 1
        request_array.append({
            "jsonrpc": "2.0",
            "id": request_id,
            "method": "eth_getTransactionCount",
            "params": [address, "latest"]
        })
    return request_array

if __name__ == '__main__':
    results = GetAccountTx(addresses)
    total_count = 0
    for row, (address, tx_count) in enumerate(zip(addresses, results), start=2):
        worksheet.cell(row=row, column=2, value=tx_count)
        total_count += tx_count
    worksheet.cell(row=len(addresses)+2, column=1, value="Total count")
    worksheet.cell(row=len(addresses)+2, column=2, value=total_count)
    workbook.save("wallets.xlsx")
